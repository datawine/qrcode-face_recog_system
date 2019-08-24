# -*- encoding: utf-8 -*-

from openpyxl import Workbook, load_workbook

def import_namelist():
    namelist = {}
    wb = load_workbook('namelist.xlsx')
    ws = wb.active
    is_first_row = True
    for row in ws.values:
        if is_first_row:
            is_first_row = False
            continue
        stu_id = str(row[1])
        person_attr = {}
        person_attr["name"] = row[0]
        person_attr["time"] = str(row[2])
        namelist[stu_id] = person_attr
    return namelist

def update_namelist(namelist):
    return namelist

def export_signin_result(signin_result):
    old_wb = load_workbook('namelist.xlsx')
    old_ws = old_wb.active
    wb = Workbook()
    ws = wb.active

    # copy ws
    row_index = 1
    for row in old_ws.values:
        col_index = 1
        for value in row:
            ws.cell(row=row_index, column=col_index).value = old_ws.cell(row=row_index, column=col_index).value
            col_index = col_index + 1
        row_index = row_index + 1

    # update time
    row_index = 0
    for row in ws.values:
        row_index = row_index + 1
        if row_index == 1:
            continue
        stu_id = str(row[1])
        if signin_result[stu_id]["time"] != "None":
            cell = ws.cell(row=row_index, column=3)
            cell.value = signin_result[stu_id]["time"]

    wb.save("output.xlsx")

if __name__ == '__main__':

    wb = Workbook()
    ws = wb.active

    namelist = import_namelist()
    print(namelist)